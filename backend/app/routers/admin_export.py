"""Export Excel mensual de solicitudes conforme Anexo 1 SEC NTCSSD."""

from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.core.auth_jwt import get_current_user
from app.database import get_db
from app.models import Solicitud, User

router = APIRouter()


_COLUMNS_FACTIBILIDAD = [
    ("NUMSOLICITUD", "num_solicitud"),
    ("FECHAINGRESO", "fecha_ingreso"),
    ("FECHAULTIMOACTUALIZACION", "fecha_ultima_actualizacion"),
    ("ESTADOACTUAL", "estado_actual"),
    ("TIPOSOLICITUD", "tipo_solicitud"),
    ("REQUIRENTE_RUT", "requirente_rut"),
    ("REQUIRENTE_NOMBRE", "requirente_nombre"),
    ("REQUIRENTE_DIRECCION", "requirente_direccion"),
    ("REQUIRENTE_TELEFONO", "requirente_telefono"),
    ("REQUIRENTE_EMAIL", "requirente_email"),
    ("PROPIETARIO_RUT", "propietario_rut"),
    ("PROPIETARIO_NOMBRE", "propietario_nombre"),
    ("PROPIETARIO_DIRECCION", "propietario_direccion"),
    ("PROPIETARIO_TELEFONO", "propietario_telefono"),
    ("PROPIETARIO_EMAIL", "propietario_email"),
    ("AUTORIZACIONTERCERO", "autorizacion_tercero"),
    ("TIPO_CONSUMO_ID", "tipo_consumo_id"),
    ("DEFINITIVOOPROVISORIO", "definitivo_provisorio"),
    ("TIPO_TENSION", "tipo_tension"),
    ("POTENCIA_SOLICITADA", "potencia_solicitada"),
    ("TIPO_FASE", "tipo_fase"),
    ("CONCESION_ID", "concesion_id"),
    ("TIPO_INSTALACION_ID", "tipo_instalacion_id"),
    ("IDENTIFICADOR_CONEXION", "identificador_conexion"),
    ("UBICACION_EMPALME", "ubicacion_empalme"),
    ("DIRECCION_INSTALACION", "direccion_instalacion"),
    ("DATUM_ID", "datum_id"),
    ("TIPO_ZONA_UTM", "tipo_zona_utm"),
    ("COORD_X", "coord_x"),
    ("COORD_Y", "coord_y"),
    ("FECHARESPUESTAEMPRESA", "fecha_respuesta_empresa"),
    ("RESPUESTAFACTIBILIDAD", "respuesta_factibilidad"),
    ("ESTUDIOS_TECNICOS_REQUERIDOS", "estudios_tecnicos_requeridos"),
    ("CAUSA_RECHAZO_NOTIFICACION", "causa_rechazo_notificacion"),
    ("OBSERVACIONES", "observaciones"),
]

_COLUMNS_CONEXION = _COLUMNS_FACTIBILIDAD + [
    ("CLIENTE_ID", "cliente_id"),
    ("MEDIDOR_ID", "medidor_id"),
    ("NUMERO_MEDIDOR", "numero_medidor"),
    ("MARCA_MEDIDOR", "marca_medidor"),
    ("MODELO_MEDIDOR", "modelo_medidor"),
    ("PUNTO_CONSUMO_ID", "punto_consumo_id"),
    ("PRESUPUESTO", "presupuesto"),
    ("DETALLE_PRESUPUESTO", "detalle_presupuesto"),
    ("MODALIDAD_FINANCIAMIENTO", "modalidad_financiamiento"),
    ("CODIGOS_VNR", "codigos_vnr"),
    ("REQUIERE_PERMISOS_TERCEROS", "requiere_permisos_terceros"),
    ("FECHA_VISITA_TERRENO", "fecha_visita_terreno"),
    ("HORA_VISITA_TERRENO", "hora_visita_terreno"),
    ("FECHA_CONEXION", "fecha_conexion"),
    ("FECHA_ENTREGA_CLIENTE_ID", "fecha_entrega_cliente_id"),
    ("OBSERVACIONES_VISITA", "observaciones_visita"),
]


def _cell_value(sol: Solicitud, attr: str):
    val = getattr(sol, attr, None)
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    return val


@router.get("/export/sec.xlsx")
async def export_sec(
    tipo: str = Query(..., pattern="^(factibilidad|conexion)$"),
    desde: date | None = None,
    hasta: date | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    if desde and hasta and desde > hasta:
        raise HTTPException(status_code=400, detail="'desde' debe ser <= 'hasta'")

    filters = [Solicitud.tipo_tramite == tipo]
    if desde:
        filters.append(Solicitud.fecha_ingreso >= desde)
    if hasta:
        filters.append(Solicitud.fecha_ingreso <= hasta)
    rows = (
        await db.execute(
            select(Solicitud).where(and_(*filters)).order_by(Solicitud.fecha_ingreso.asc())
        )
    ).scalars().all()

    columns = _COLUMNS_FACTIBILIDAD if tipo == "factibilidad" else _COLUMNS_CONEXION

    wb = Workbook()
    ws = wb.active
    ws.title = "factibilidad" if tipo == "factibilidad" else "conexion"

    header_font = Font(bold=True, color="111111")
    header_fill = PatternFill("solid", fgColor="EFA829")
    for col_idx, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, sol in enumerate(rows, start=2):
        for col_idx, (_, attr) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(sol, attr))

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%d%m%Y")
    fname = f"{settings.EMPRESA_CODIGO}_{today}_{tipo}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
